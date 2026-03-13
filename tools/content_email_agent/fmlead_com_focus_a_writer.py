import argparse
import base64
import datetime as dt
import email
import html
import imaplib
import json
import os
import re
import socket
import smtplib
import shutil
import subprocess
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from dataclasses import dataclass
from email.message import EmailMessage
from pathlib import Path
from typing import Dict, Optional

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook

DRAFT_TOKEN_RE = re.compile(r"\[FM-LEAD-DRAFT:([a-z0-9\-]+)\]", re.IGNORECASE)
TOPIC_SUBJECT_PREFIX = "Content Approval Request- Focus Mindset - Articles - /"
TOPIC_ID_RE = re.compile(r"topic\s*id\s*:\s*([a-z0-9]{8,32})", re.IGNORECASE)
WATCH_SINGLETON_PORT = 47291
AGENT_NAME = "FMLead.com_Focus_A_Writer"


@dataclass
class Config:
    openai_api_key: str
    openai_model: str
    smtp_host: str
    smtp_port: int
    smtp_username: str
    smtp_password: str
    imap_host: str
    imap_port: int
    imap_username: str
    imap_password: str
    owner_email: str
    hugo_site_path: Path
    article_root_relative: str
    default_tags: str
    poll_seconds: int
    image_model: str
    image_size: str
    image_quality: str
    image_style: str
    image_prompt_suffix: str
    image_provider: str
    image_query_mode: str
    generate_detail_image: bool
    in_post_image_count: int
    llm_backend: str
    ollama_url: str
    ollama_model: str
    ollama_timeout_seconds: int
    max_source_chars: int
    enable_url_fetch: bool
    generate_images: bool
    tracker_filename: str
    daily_topics_enabled: bool
    daily_topics_count: int
    daily_topics_hour_24: int
    idea_pipeline_max: int
    weekly_idea_reminder_enabled: bool
    weekly_reminder_day_0_mon: int
    weekly_reminder_hour_24: int
    attach_images_in_draft_email: bool
    image_thumb_filename: str
    image_body1_filename: str
    image_body2_filename: str


class FMLeadComFocusAWriter:
    def __init__(self, config: Config, state_path: Path):
        self.config = config
        self.state_path = state_path
        self.client = OpenAI(api_key=self.config.openai_api_key) if self.config.openai_api_key else None
        self.state = self._load_state()
        self._watch_lock_socket: Optional[socket.socket] = None

    def _load_state(self) -> Dict:
        if self.state_path.exists():
            state = json.loads(self.state_path.read_text(encoding="utf-8"))
            state.setdefault("drafts", {})
            state.setdefault("processed_message_ids", [])
            state.setdefault("topics", {})
            state.setdefault("last_daily_topics_date", "")
            state.setdefault("last_weekly_topic_reminder_week", "")
            return state
        return {
            "drafts": {},
            "processed_message_ids": [],
            "topics": {},
            "last_daily_topics_date": "",
            "last_weekly_topic_reminder_week": "",
        }

    def _save_state(self) -> None:
        self.state_path.write_text(json.dumps(self.state, indent=2), encoding="utf-8")

    def _timing_log_path(self) -> Path:
        return self.state_path.parent / "timing.log"

    def _append_timing_log(self, event: str, payload: Dict[str, object]) -> None:
        line = {
            "ts": dt.datetime.now(dt.timezone.utc).isoformat(),
            "event": event,
            **payload,
        }
        try:
            with self._timing_log_path().open("a", encoding="utf-8") as fp:
                fp.write(json.dumps(line, ensure_ascii=True) + "\n")
        except Exception as exc:
            print(f"timing log write failed: {exc}")

    def _article_root(self) -> Path:
        return self.config.hugo_site_path / self.config.article_root_relative

    def _tracker_path(self) -> Path:
        return self.state_path.parent / self.config.tracker_filename

    def _sources_root(self) -> Path:
        root = self.state_path.parent / "inbox_sources"
        root.mkdir(parents=True, exist_ok=True)
        return root

    def _reply_images_root(self) -> Path:
        root = self.state_path.parent / "reply_images"
        root.mkdir(parents=True, exist_ok=True)
        return root

    def _reply_images_dir(self, draft_id: str) -> Path:
        draft = self.state.get("drafts", {}).get(draft_id, {})
        slug = self._slugify(str(draft.get("slug") or draft.get("title") or draft_id)) or draft_id
        folder = self._reply_images_root() / slug
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def _slugify(self, text: str) -> str:
        cleaned = re.sub(r"[^a-zA-Z0-9\s-]", "", text).strip().lower()
        return re.sub(r"[\s-]+", "-", cleaned).strip("-")

    def _approval_subject(self, title: str) -> str:
        return f"{TOPIC_SUBJECT_PREFIX}{title.strip()}"

    def _normalize_space(self, value: str) -> str:
        return re.sub(r"\s+", " ", (value or "")).strip()

    def _extract_title_from_topic_subject(self, subject: str) -> str:
        normalized_subject = self._normalize_space(subject)
        lower = normalized_subject.lower()
        prefix = TOPIC_SUBJECT_PREFIX.lower()
        idx = lower.find(prefix)
        if idx < 0:
            return ""
        raw = normalized_subject[idx + len(TOPIC_SUBJECT_PREFIX) :]
        return self._normalize_space(raw)

    def _find_topic_by_title(self, title: str) -> Optional[Dict[str, str]]:
        target = self._normalize_space(title).lower()
        if not target:
            return None
        for topic in self.state.get("topics", {}).values():
            candidate = self._normalize_space(str(topic.get("title", ""))).lower()
            if candidate == target:
                return topic
        return None

    def _find_topic_by_feedback_text(self, feedback_text: str) -> Optional[Dict[str, str]]:
        match = TOPIC_ID_RE.search(feedback_text or "")
        if not match:
            return None
        topic_id = match.group(1).lower()
        topic = self.state.get("topics", {}).get(topic_id)
        return topic if isinstance(topic, dict) else None

    def _is_rejected(self, cleaned_feedback: str) -> bool:
        for line in cleaned_feedback.splitlines():
            if line.strip().lower() == "rejected":
                return True
        return False

    def _open_topic_count(self) -> int:
        topics = self.state.get("topics", {})
        open_statuses = {"awaiting_approval", "idea"}
        return sum(1 for topic in topics.values() if str(topic.get("status", "")).strip().lower() in open_statuses)

    def _topics_awaiting_reply(self) -> list[Dict[str, str]]:
        topics = self.state.get("topics", {})
        pending = [
            topic
            for topic in topics.values()
            if str(topic.get("status", "")).strip().lower() == "awaiting_approval"
        ]
        pending.sort(key=lambda t: str(t.get("created_at", "")))
        return pending

    def _weekly_reminder_week_key(self, when: dt.datetime) -> str:
        iso = when.isocalendar()
        return f"{iso.year}-W{iso.week:02d}"

    def _build_weekly_topic_reminder_body(self, pending_topics: list[Dict[str, str]], now: dt.datetime) -> str:
        lines = [
            "Weekly reminder: pending Focus Mindset topic approvals.",
            "Reply to any original topic email with exactly 'approved' or 'rejected'.",
            "",
            f"Pending topics: {len(pending_topics)}",
            "",
        ]
        for idx, topic in enumerate(pending_topics, start=1):
            title = str(topic.get("title", "(untitled)"))
            topic_id = str(topic.get("id", ""))
            created_raw = str(topic.get("created_at", ""))
            age_days_text = ""
            try:
                created_dt = dt.datetime.fromisoformat(created_raw.replace("Z", "+00:00"))
                now_utc = dt.datetime.now(dt.timezone.utc)
                age_days = max(0, (now_utc - created_dt).days)
                age_days_text = f" ({age_days} days old)"
            except Exception:
                pass
            lines.append(f"{idx}. {title}{age_days_text}")
            lines.append(f"   Topic ID: {topic_id}")

        return "\n".join(lines)

    def _handle_topic_feedback(self, topic: Dict[str, str], feedback: str) -> None:
        cleaned = self._normalize_feedback(feedback)
        if not cleaned:
            return

        if self._is_approved(cleaned):
            topic["status"] = "approved"
            topic["approved_at"] = dt.datetime.utcnow().isoformat() + "Z"
            self._save_state()

            started = time.time()
            draft_id = self.create_new_draft(title=str(topic.get("title", "")), brief=str(topic.get("brief", "")))
            elapsed_seconds = round(time.time() - started, 2)

            topic["status"] = "draft_created"
            topic["draft_id"] = draft_id
            self._save_state()
            self._append_timing_log(
                event="topic_approved_to_draft_email",
                payload={
                    "topic_id": str(topic.get("id", "")),
                    "draft_id": draft_id,
                    "title": str(topic.get("title", "")),
                    "elapsed_seconds": elapsed_seconds,
                },
            )
            print(
                f"timing: topic approved -> draft emailed in {elapsed_seconds}s "
                f"(topic_id={topic.get('id','')}, draft_id={draft_id})"
            )
            return

        if self._is_rejected(cleaned):
            topic["status"] = "rejected"
            topic["rejected_at"] = dt.datetime.utcnow().isoformat() + "Z"
            topic["last_feedback"] = cleaned[:1000]
            self._save_state()
            return

        topic["status"] = "idea"
        topic["last_feedback"] = cleaned[:1000]
        self._save_state()

    def _ai_write(self, prompt: str) -> str:
        if self.config.llm_backend == "ollama":
            return self._ollama_write(prompt)

        if self.client is None:
            raise ValueError("OpenAI backend selected but OPENAI_API_KEY is empty")

        response = self.client.responses.create(
            model=self.config.openai_model,
            input=[
                {
                    "role": "system",
                    "content": (
                        "You write practical, no-fluff FM-Lead Focus Mindset articles. "
                        "Return markdown only. Do not include YAML front matter."
                    ),
                },
                {"role": "user", "content": prompt},
            ],
        )
        return response.output_text.strip()

    def _ollama_write(self, prompt: str) -> str:
        payload = {
            "model": self.config.ollama_model,
            "prompt": prompt,
            "stream": False,
        }
        req = urllib.request.Request(
            url=f"{self.config.ollama_url.rstrip('/')}/api/generate",
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=self.config.ollama_timeout_seconds) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="replace"))
        return str(data.get("response", "")).strip()

    def _build_draft_prompt(self, title: str, brief: str) -> str:
        return (
            f"Write a Focus Mindset article titled '{title}'.\n"
            f"Brief: {brief}\n"
            "Audience: facility management and critical operations leaders.\n"
            "Length: 700-1100 words.\n"
            "Style: practical, reflective, clear, and actionable.\n"
            "Use headings and short sections. End with a short action checklist."
        )

    def _build_topic_ideas_prompt(self, count: int) -> str:
        return (
            "Generate topic ideas for FM-Lead Focus Mindset articles. "
            f"Return JSON only as an array of {count} objects with keys: title, brief. "
            "Topics must be practical for facilities managers and critical operations leaders."
        )

    def _image_reminder_block(self) -> str:
        return (
            "Approval actions (reply in this same thread):\n"
            "- To publish now: include one line exactly: approved\n"
            "- To decline this idea/draft: include one line exactly: rejected\n"
            "- To request edits: reply with your edit notes (do not use approved/rejected)\n\n"
            "If you want to provide your own images, attach JPG/JPEG/PNG files using strict filename labels:\n"
            "- thumb_*  -> featured image\n"
            "- body1_*  -> first in-post image (after intro)\n"
            "- body2_*  -> second in-post image (before conclusion)\n\n"
            "Optional image labels in reply body (used as alt text/captions):\n"
            "- thumb label: <text>\n"
            "- body1 label: <text>\n"
            "- body2 label: <text>\n\n"
            "Example reply:\n"
            "approved\n"
            "thumb label: Exterior of facility operations center\n"
            "body1 label: Technician inspecting HVAC panel\n"
            "body2 label: CMMS dashboard review with team\n\n"
            "Important: if attached image filenames do not include thumb/body1/body2, publishing is paused until corrected."
        )

    def _with_image_reminder(self, body: str) -> str:
        return f"{body.rstrip()}\n\n---\n\n{self._image_reminder_block()}"

    def _generate_topic_ideas(self, count: int) -> list[Dict[str, str]]:
        fallback = [
            {
                "title": "Leading Calm During Facility Incidents",
                "brief": "Practical leadership habits for keeping teams focused during disruptions.",
            },
            {
                "title": "Weekly Focus Cadence for FM Teams",
                "brief": "A repeatable weekly execution rhythm for maintenance and operations.",
            },
            {
                "title": "Reducing Task Switching in Control Rooms",
                "brief": "How to design workdays that reduce context switching and improve reliability.",
            },
            {
                "title": "CMMS Discipline That Improves Execution",
                "brief": "How focused CMMS hygiene supports accountability and uptime outcomes.",
            },
            {
                "title": "Decision Checklists for Facilities Leaders",
                "brief": "Simple checklists to improve judgment and consistency under pressure.",
            },
        ]
        try:
            raw = self._ai_write(self._build_topic_ideas_prompt(count=count)).strip()
            match = re.search(r"\[[\s\S]*\]", raw)
            candidate = match.group(0) if match else raw
            parsed = json.loads(candidate)
            if not isinstance(parsed, list):
                return fallback[:count]
            topics = []
            for item in parsed:
                title = str(item.get("title", "")).strip()
                brief = str(item.get("brief", "")).strip()
                if title and brief:
                    topics.append({"title": title, "brief": brief})
            return topics[:count] if topics else fallback[:count]
        except Exception as exc:
            print(f"topic generation failed: {exc}")
            return fallback[:count]

    def send_daily_topic_approval_requests(self) -> int:
        open_count = self._open_topic_count()
        available_slots = max(0, self.config.idea_pipeline_max - open_count)
        if available_slots <= 0:
            return 0

        request_count = min(self.config.daily_topics_count, available_slots)
        ideas = self._generate_topic_ideas(count=request_count)
        sent = 0
        for idea in ideas:
            topic_id = uuid.uuid4().hex[:10]
            title = idea["title"]
            brief = idea["brief"]
            self.state["topics"][topic_id] = {
                "id": topic_id,
                "title": title,
                "brief": brief,
                "status": "awaiting_approval",
                "created_at": dt.datetime.utcnow().isoformat() + "Z",
            }
            self._save_state()

            subject = self._approval_subject(title)
            body = (
                "Daily content approval request for Focus Mindset articles.\n\n"
                f"Topic ID: {topic_id}\n"
                f"Title: {title}\n"
                f"Brief: {brief}\n\n"
                "Reply with a line exactly: approved\n"
                "Or reply with a line exactly: rejected\n"
                "Or reply with revision instructions for this idea."
            )
            self._send_email(subject=subject, body=self._with_image_reminder(body))
            sent += 1
        return sent

    def _maybe_send_weekly_topic_reminder(self) -> None:
        if not self.config.weekly_idea_reminder_enabled:
            return

        now = dt.datetime.now()
        if now.weekday() != self.config.weekly_reminder_day_0_mon:
            return
        if now.hour < self.config.weekly_reminder_hour_24:
            return

        week_key = self._weekly_reminder_week_key(now)
        if self.state.get("last_weekly_topic_reminder_week") == week_key:
            return

        pending_topics = self._topics_awaiting_reply()
        if not pending_topics:
            self.state["last_weekly_topic_reminder_week"] = week_key
            self._save_state()
            return

        subject = "Weekly Reminder - Pending FM-Lead Topic Approvals"
        body = self._build_weekly_topic_reminder_body(pending_topics=pending_topics, now=now)
        self._send_email(subject=subject, body=self._with_image_reminder(body))
        self.state["last_weekly_topic_reminder_week"] = week_key
        self._save_state()

    def _maybe_send_daily_topics(self) -> None:
        if not self.config.daily_topics_enabled:
            return

        now = dt.datetime.now()
        today = now.strftime("%Y-%m-%d")
        if now.hour < self.config.daily_topics_hour_24:
            return
        if self.state.get("last_daily_topics_date") == today:
            return

        sent = self.send_daily_topic_approval_requests()
        if sent > 0:
            self.state["last_daily_topics_date"] = today
            self._save_state()

    def _build_revision_prompt(self, current_markdown: str, feedback: str, source_context: str) -> str:
        source_block = (
            "\nAdditional source material from email attachments/URLs:\n"
            f"{source_context}\n"
            if source_context.strip()
            else ""
        )
        return (
            "Revise the article based on the editor feedback.\n"
            "Keep article quality and flow strong.\n"
            "Use source material for factual grounding and include inline citations like [1], [2].\n"
            "Return full revised markdown only.\n"
            f"Feedback:\n{feedback}\n\n"
            f"{source_block}\n"
            f"Current article:\n{current_markdown}"
        )

    def _topic_image_directive(self, title: str, brief: str) -> str:
        text = f"{title} {brief}".lower()
        if any(k in text for k in ["resilien", "recovery", "setback", "bounce back", "adapt"]):
            return "resilience: subtle storm-to-calm contrast, recovery and composure"
        if any(k in text for k in ["leader", "manag", "team", "mentor", "culture"]):
            return "leadership: collaborative setting, confident guidance, human-centered"
        if any(k in text for k in ["execut", "discipline", "process", "ops", "operation", "procedure"]):
            return "execution: checklists, precision workflow, operational rigor"
        if any(k in text for k in ["focus", "attention", "deep work", "concentration"]):
            return "focus: minimal distractions, single-point attention, calm intensity"
        return "general: professional editorial mood aligned to focus mindset"

    def _build_image_prompt(self, title: str, brief: str) -> str:
        topic_directive = self._topic_image_directive(title=title, brief=brief)
        base_prompt = (
            f"Create a clean, professional hero image for an article titled '{title}'. "
            "Theme: focus mindset, disciplined execution, calm leadership under pressure. "
            "Visual style: editorial, modern, realistic lighting, no logos, no text overlays. "
            "Avoid clutter. Keep composition wide and suitable for website feature image. "
            f"Context: {brief}. Topic emphasis: {topic_directive}."
        )
        suffix = self.config.image_prompt_suffix.strip()
        return f"{base_prompt} {suffix}".strip() if suffix else base_prompt

    def _build_detail_image_prompt(self, title: str, brief: str) -> str:
        topic_directive = self._topic_image_directive(title=title, brief=brief)
        base_prompt = (
            f"Create a supporting editorial image for an article titled '{title}'. "
            "Use a different composition from the hero image and focus on a concrete execution moment. "
            "No logos, no text overlays, no watermarks. "
            f"Topic emphasis: {topic_directive}. Context: {brief}."
        )
        suffix = self.config.image_prompt_suffix.strip()
        return f"{base_prompt} {suffix}".strip() if suffix else base_prompt

    def _write_image_from_response(self, response, image_file: Path) -> Optional[str]:
        data = getattr(response, "data", None) or []
        if not data:
            return None

        first = data[0]
        b64_json = getattr(first, "b64_json", None)
        url = getattr(first, "url", None)
        if isinstance(first, dict):
            b64_json = b64_json or first.get("b64_json")
            url = url or first.get("url")

        if b64_json:
            image_file.write_bytes(base64.b64decode(b64_json))
            return image_file.name

        if url:
            try:
                with urllib.request.urlopen(url, timeout=60) as resp:
                    image_file.write_bytes(resp.read())
                return image_file.name
            except urllib.error.URLError as exc:
                print(f"image download failed: {exc}")
                return None

        return None

    def _download_image(self, url: str, image_file: Path, retries: int = 2) -> Optional[str]:
        last_exc = None
        for attempt in range(retries + 1):
            try:
                req = urllib.request.Request(
                    url=url,
                    headers={"User-Agent": "fmlead-content-agent/1.0"},
                    method="GET",
                )
                with urllib.request.urlopen(req, timeout=90) as resp:
                    image_file.write_bytes(resp.read())
                return image_file.name
            except Exception as exc:
                last_exc = exc
                # Small backoff between retries.
                time.sleep(1 + attempt)

        print(f"image download failed after retries: {last_exc}")
        return None

    def _unsplash_query_for_kind(self, kind: str, title: str, brief: str) -> str:
        base = "facility management operations focus"
        if kind == "feature":
            return f"{base} leadership control room"
        if kind == "detail-1":
            return f"{base} maintenance technician inspection"
        if kind == "detail-2":
            return f"{base} cmms dashboard planning team"
        return f"{base} {title} {brief}"

    def _kind_query_tail(self, kind: str) -> str:
        if kind == "feature":
            return "strategic leadership calm decision making"
        if kind == "detail-1":
            return "hands-on maintenance inspection on site"
        if kind == "detail-2":
            return "planning board dashboard team coordination"
        return "facility operations"

    def _build_distinct_image_query(self, kind: str, title: str, brief: str, suggested_query: Optional[str]) -> str:
        root = (suggested_query or self._unsplash_query_for_kind(kind=kind, title=title, brief=brief)).strip()
        # Keep key terms from title/brief to preserve relevance if model suggestions are generic.
        title_terms = " ".join(re.findall(r"[A-Za-z0-9]+", title)[:4])
        brief_terms = " ".join(re.findall(r"[A-Za-z0-9]+", brief)[:4])
        return f"{root} {title_terms} {brief_terms} {self._kind_query_tail(kind)}".strip()

    def _suggest_image_queries(self, title: str, brief: str, markdown: str) -> Dict[str, str]:
        article_text = re.sub(r"\s+", " ", markdown)
        article_text = article_text[:4000]
        prompt = (
            "You are helping select highly relevant editorial stock-photo search terms for a facilities management article. "
            "Return JSON only with keys: feature, detail1, detail2. "
            "Each value should be a short search phrase (4-8 words), concrete, no quotes, no punctuation.\n\n"
            f"Title: {title}\n"
            f"Brief: {brief}\n"
            f"Article excerpt: {article_text}"
        )

        fallback = {
            "feature": self._unsplash_query_for_kind("feature", title, brief),
            "detail1": self._unsplash_query_for_kind("detail-1", title, brief),
            "detail2": self._unsplash_query_for_kind("detail-2", title, brief),
        }

        try:
            raw = self._ai_write(prompt).strip()
            match = re.search(r"\{[\s\S]*\}", raw)
            candidate = match.group(0) if match else raw
            parsed = json.loads(candidate)
            return {
                "feature": str(parsed.get("feature", fallback["feature"]))[:120],
                "detail1": str(parsed.get("detail1", fallback["detail1"]))[:120],
                "detail2": str(parsed.get("detail2", fallback["detail2"]))[:120],
            }
        except Exception as exc:
            print(f"image query suggestion failed: {exc}")
            return fallback

    def _generate_unsplash_image(
        self,
        kind: str,
        title: str,
        brief: str,
        image_file: Path,
        seed: str,
        suggested_query: Optional[str] = None,
    ) -> Optional[str]:
        query = self._build_distinct_image_query(
            kind=kind,
            title=title,
            brief=brief,
            suggested_query=suggested_query,
        )
        encoded = urllib.parse.quote_plus(query)
        seed_num = abs(hash(seed)) % 10000
        tags = re.findall(r"[a-z0-9]+", query.lower())[:6]
        tag_csv = ",".join(tags) if tags else "facility,management,operations"
        cache_bust = int(time.time() * 1000)

        candidate_urls = [
            # Primary free source with cache-busting to reduce duplicate returns.
            f"https://source.unsplash.com/featured/1600x900/?{encoded}&sig={seed_num}&cb={cache_bust}",
            # Fallback 1: keyword-based image with properly formatted CSV tags.
            f"https://loremflickr.com/1600/900/{tag_csv}/all?lock={seed_num}",
            # Fallback 2: deterministic but distinct placeholder per slot.
            f"https://picsum.photos/seed/{seed_num}-{kind}/1600/900",
        ]

        for url in candidate_urls:
            saved = self._download_image(url=url, image_file=image_file, retries=1)
            if saved:
                return saved

        return None

    def _generate_openai_image(self, prompt: str, image_file: Path) -> Optional[str]:
        if self.client is None:
            return None
        try:
            response = self.client.images.generate(
                model=self.config.image_model,
                prompt=prompt,
                size=self.config.image_size,
                quality=self.config.image_quality,
                style=self.config.image_style,
            )
        except Exception as exc:
            print(f"image generation failed: {exc}")
            return None
        return self._write_image_from_response(response=response, image_file=image_file)

    def _generate_feature_image(
        self,
        title: str,
        brief: str,
        images_dir: Path,
        suggested_query: Optional[str] = None,
    ) -> Optional[str]:
        if not self.config.generate_images:
            return None

        image_file = images_dir / "feature.png"
        provider = self.config.image_provider.lower()
        if provider == "openai":
            prompt = self._build_image_prompt(title=title, brief=brief)
            return self._generate_openai_image(prompt=prompt, image_file=image_file)
        if provider == "unsplash":
            return self._generate_unsplash_image(
                kind="feature",
                title=title,
                brief=brief,
                image_file=image_file,
                seed=f"{self._slugify(title)}-feature",
                suggested_query=suggested_query,
            )
        return None

    def _generate_inpost_images(
        self,
        title: str,
        brief: str,
        images_dir: Path,
        suggested_queries: Optional[Dict[str, str]] = None,
    ) -> list[str]:
        if not self.config.generate_images:
            return []
        if not self.config.generate_detail_image:
            return []

        count = max(1, min(self.config.in_post_image_count, 2))
        names = []
        provider = self.config.image_provider.lower()

        for idx in range(1, count + 1):
            kind = f"detail-{idx}"
            image_file = images_dir / f"detail-{idx}.png"
            if provider == "openai":
                prompt = self._build_detail_image_prompt(title=title, brief=brief)
                saved = self._generate_openai_image(prompt=prompt, image_file=image_file)
            elif provider == "unsplash":
                key = "detail1" if idx == 1 else "detail2"
                saved = self._generate_unsplash_image(
                    kind=kind,
                    title=title,
                    brief=brief,
                    image_file=image_file,
                    seed=f"{self._slugify(title)}-{kind}",
                    suggested_query=(suggested_queries or {}).get(key),
                )
            else:
                saved = None

            if saved:
                names.append(saved)

        return names

    def _send_email(self, subject: str, body: str, attachments: Optional[list[Path]] = None) -> None:
        msg = EmailMessage()
        msg["From"] = self.config.smtp_username
        msg["To"] = self.config.owner_email
        msg["Subject"] = subject
        msg["X-FMLead-Agent"] = "1"
        tracker_path = self._update_tracker_workbook()
        msg.set_content(body + "\n\nAttached: FM-Lead post status tracker (Excel).")

        if tracker_path.exists():
            msg.add_attachment(
                tracker_path.read_bytes(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=tracker_path.name,
            )

        for path in attachments or []:
            if not path.exists():
                continue
            msg.add_attachment(
                path.read_bytes(),
                maintype="image",
                subtype="png",
                filename=path.name,
            )

        with smtplib.SMTP_SSL(self.config.smtp_host, self.config.smtp_port) as server:
            server.login(self.config.smtp_username, self.config.smtp_password)
            server.send_message(msg)

    def _draft_images_dir(self, draft_id: str) -> Path:
        d = self.state_path.parent / "draft_images" / draft_id
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _create_draft_email_images(self, draft_id: str, title: str, brief: str, markdown: str) -> list[Path]:
        if not self.config.generate_images or not self.config.attach_images_in_draft_email:
            return []
        images_dir = self._draft_images_dir(draft_id)
        image_queries: Dict[str, str] = {}
        if self.config.image_query_mode == "source":
            image_queries = self._suggest_image_queries(title=title, brief=brief, markdown=markdown)

        created: list[Path] = []
        feature = self._generate_feature_image(
            title=title,
            brief=brief,
            images_dir=images_dir,
            suggested_query=image_queries.get("feature"),
        )
        if feature:
            created.append(images_dir / feature)

        details = self._generate_inpost_images(
            title=title,
            brief=brief,
            images_dir=images_dir,
            suggested_queries=image_queries,
        )
        for name in details:
            created.append(images_dir / name)

        return created

    def _parse_frontmatter(self, content: str) -> Dict[str, str]:
        if not content.startswith("---"):
            return {}
        lines = content.splitlines()
        if len(lines) < 3:
            return {}

        frontmatter_lines = []
        for line in lines[1:]:
            if line.strip() == "---":
                break
            frontmatter_lines.append(line)

        parsed: Dict[str, str] = {}
        for line in frontmatter_lines:
            if ":" not in line:
                continue
            key, value = line.split(":", 1)
            parsed[key.strip().lower()] = value.strip().strip('"').strip("'")
        return parsed

    def _state_status_label(self, status: str, markdown: str) -> str:
        normalized = status.strip().lower()
        if normalized == "published":
            return "published"
        if normalized == "idea" or not (markdown or "").strip():
            return "idea"
        return "draft mode"

    def _content_rows(self) -> list[Dict[str, str]]:
        rows: list[Dict[str, str]] = []
        article_root = self._article_root()
        if not article_root.exists():
            return rows

        for child in sorted(article_root.iterdir()):
            if child.name.startswith("_"):
                continue

            index_path = child / "index.md" if child.is_dir() else child
            if not index_path.exists() or index_path.suffix.lower() != ".md":
                continue

            try:
                text = index_path.read_text(encoding="utf-8")
            except Exception:
                continue

            fm = self._parse_frontmatter(text)
            title = fm.get("title") or child.stem.replace("-", " ").title()
            draft_value = fm.get("draft", "false").lower()
            status = "draft mode" if draft_value in {"true", "yes", "1"} else "published"
            post_date = fm.get("date", "")

            rows.append(
                {
                    "title": title,
                    "slug": child.name if child.is_dir() else child.stem,
                    "status": status,
                    "date": post_date,
                    "source": "content directory",
                    "path": str(index_path.relative_to(self.config.hugo_site_path)).replace("\\", "/"),
                    "state_id": "",
                }
            )

        return rows

    def _state_rows(self) -> list[Dict[str, str]]:
        rows: list[Dict[str, str]] = []
        drafts = self.state.get("drafts", {})
        for draft_id, draft in drafts.items():
            slug = str(draft.get("slug", "")).strip()
            title = str(draft.get("title", "")).strip() or slug
            status = self._state_status_label(str(draft.get("status", "")), str(draft.get("markdown", "")))
            rows.append(
                {
                    "title": title,
                    "slug": slug,
                    "status": status,
                    "date": str(draft.get("created_at", ""))[:10],
                    "source": "email agent state",
                    "path": f"{self.config.article_root_relative}/{slug}/index.md" if slug else "",
                    "state_id": draft_id,
                }
            )
        return rows

    def _update_tracker_workbook(self) -> Path:
        tracker_path = self._tracker_path()
        wb = Workbook()
        ws = wb.active
        ws.title = "Post Tracker"
        ws.append(["Title", "Slug", "Status", "Date", "Source", "Path", "State ID"])

        merged: Dict[str, Dict[str, str]] = {}
        for row in self._content_rows():
            merged[row["slug"]] = row

        for row in self._state_rows():
            key = row["slug"] or row["state_id"]
            existing = merged.get(key)
            if existing is None:
                merged[key] = row
                continue

            # Preserve published status from content unless state says published explicitly.
            if existing["status"] == "published" and row["status"] != "published":
                continue
            merged[key] = row

        status_order = {"published": 0, "draft mode": 1, "idea": 2}
        sorted_rows = sorted(
            merged.values(),
            key=lambda r: (status_order.get(r.get("status", "idea"), 9), r.get("title", "").lower()),
        )

        for row in sorted_rows:
            ws.append([
                row.get("title", ""),
                row.get("slug", ""),
                row.get("status", ""),
                row.get("date", ""),
                row.get("source", ""),
                row.get("path", ""),
                row.get("state_id", ""),
            ])

        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 24

        tracker_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(tracker_path)
        return tracker_path

    def _extract_text_body(self, msg: email.message.Message) -> str:
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain" and "attachment" not in str(part.get("Content-Disposition", "")):
                    payload = part.get_payload(decode=True)
                    if payload is None:
                        continue
                    charset = part.get_content_charset() or "utf-8"
                    return payload.decode(charset, errors="replace")
            return ""
        payload = msg.get_payload(decode=True)
        if payload is None:
            return ""
        charset = msg.get_content_charset() or "utf-8"
        return payload.decode(charset, errors="replace")

    def _decode_bytes(self, data: bytes) -> str:
        for enc in ("utf-8", "latin-1"):
            try:
                return data.decode(enc)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    def _extract_urls(self, text: str) -> list[str]:
        urls = re.findall(r"https?://[^\s)\]>]+", text)
        return list(dict.fromkeys(urls))

    def _strip_html(self, html_text: str) -> str:
        no_script = re.sub(r"<script[\\s\\S]*?</script>", " ", html_text, flags=re.IGNORECASE)
        no_style = re.sub(r"<style[\\s\\S]*?</style>", " ", no_script, flags=re.IGNORECASE)
        text = re.sub(r"<[^>]+>", " ", no_style)
        text = html.unescape(text)
        return re.sub(r"\s+", " ", text).strip()

    def _fetch_url_text(self, url: str) -> str:
        if not self.config.enable_url_fetch:
            return ""
        try:
            req = urllib.request.Request(
                url=url,
                headers={"User-Agent": "fmlead-content-agent/1.0"},
                method="GET",
            )
            with urllib.request.urlopen(req, timeout=45) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
            return self._strip_html(raw)[: self.config.max_source_chars]
        except Exception as exc:
            return f"Could not fetch URL {url}: {exc}"

    def _extract_text_from_attachment(self, filename: str, data: bytes) -> str:
        ext = Path(filename).suffix.lower()
        if ext in {".png", ".jpg", ".jpeg"}:
            # Image files are handled separately for publishing and are not used as text sources.
            return ""
        if ext in {".txt", ".md", ".csv", ".json"}:
            return self._decode_bytes(data)
        if ext == ".pdf":
            try:
                import io
                from pypdf import PdfReader

                reader = PdfReader(io.BytesIO(data))
                return "\n".join([page.extract_text() or "" for page in reader.pages])
            except Exception as exc:
                return f"Could not parse PDF {filename}: {exc}"
        if ext == ".docx":
            try:
                import io
                from docx import Document

                doc = Document(io.BytesIO(data))
                return "\n".join([p.text for p in doc.paragraphs if p.text])
            except Exception as exc:
                return f"Could not parse DOCX {filename}: {exc}"
        if ext in {".xlsx", ".xlsm", ".xltx"}:
            try:
                import io
                from openpyxl import load_workbook

                wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                lines = []
                for sheet in wb.worksheets:
                    lines.append(f"Sheet: {sheet.title}")
                    for row in sheet.iter_rows(values_only=True):
                        vals = [str(v) for v in row if v is not None and str(v).strip()]
                        if vals:
                            lines.append(" | ".join(vals))
                return "\n".join(lines)
            except Exception as exc:
                return f"Could not parse Excel {filename}: {exc}"
        return f"Unsupported attachment type: {filename}"

    def _classify_image_role(self, filename: str) -> Optional[str]:
        name = filename.lower()
        if "thumb" in name or "thumbnail" in name or "feature" in name:
            return "thumb"
        if "body1" in name or "detail-1" in name or "detail1" in name or "inpost1" in name or "in-post-1" in name:
            return "body1"
        if "body2" in name or "detail-2" in name or "detail2" in name or "inpost2" in name or "in-post-2" in name:
            return "body2"
        return None

    def _extract_image_labels_from_feedback(self, feedback: str) -> Dict[str, str]:
        labels: Dict[str, str] = {}
        patterns = {
            "thumb": [r"^\s*(?:thumb|thumbnail|feature)\s*label\s*:\s*(.+)$"],
            "body1": [r"^\s*(?:body1|inbody1|in-body-1|detail1)\s*label\s*:\s*(.+)$"],
            "body2": [r"^\s*(?:body2|inbody2|in-body-2|detail2)\s*label\s*:\s*(.+)$"],
        }

        for key, regexes in patterns.items():
            for rgx in regexes:
                match = re.search(rgx, feedback, flags=re.IGNORECASE | re.MULTILINE)
                if match:
                    value = re.sub(r"\s+", " ", match.group(1)).strip()
                    if value:
                        labels[key] = value[:180]
                        break
        return labels

    def _extract_feedback_sources(self, msg: email.message.Message, draft_id: str) -> tuple[str, str, list[Dict[str, str]]]:
        body_text = self._extract_text_body(msg).strip()
        sources = []
        user_images: list[Dict[str, str]] = []
        draft_folder = self._sources_root() / draft_id
        draft_folder.mkdir(parents=True, exist_ok=True)
        image_folder = self._reply_images_dir(draft_id=draft_id)

        for part in msg.walk():
            disposition = str(part.get("Content-Disposition", ""))
            filename = part.get_filename()
            if not filename or "attachment" not in disposition.lower():
                continue
            payload = part.get_payload(decode=True)
            if payload is None:
                continue

            safe_name = Path(filename).name
            stamp = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            saved_path = draft_folder / f"{stamp}_{safe_name}"
            saved_path.write_bytes(payload)

            if Path(safe_name).suffix.lower() in {".png", ".jpg", ".jpeg"}:
                img_saved = image_folder / f"{stamp}_{safe_name}"
                img_saved.write_bytes(payload)
                user_images.append(
                    {
                        "path": str(img_saved),
                        "filename": safe_name,
                        "role": self._classify_image_role(safe_name) or "",
                        "received_at": dt.datetime.utcnow().isoformat() + "Z",
                    }
                )
                continue

            extracted = self._extract_text_from_attachment(safe_name, payload)
            if extracted.strip():
                sources.append(f"Attachment: {safe_name}\n{extracted[: self.config.max_source_chars]}")

        for url in self._extract_urls(body_text):
            parsed = urllib.parse.urlparse(url)
            if parsed.scheme in {"http", "https"}:
                fetched = self._fetch_url_text(url)
                if fetched:
                    sources.append(f"URL: {url}\n{fetched}")

        return body_text, "\n\n".join(sources), user_images

    def _select_user_publish_images(self, draft: Dict, images_dir: Path) -> tuple[Optional[str], list[str]]:
        raw_items = draft.get("user_image_attachments", [])
        if not isinstance(raw_items, list) or not raw_items:
            return None, []

        candidates = []
        for item in raw_items:
            if isinstance(item, dict):
                path = Path(str(item.get("path", "")))
                role = str(item.get("role", "")).strip().lower()
            else:
                path = Path(str(item))
                role = self._classify_image_role(path.name) or ""

            if not path.exists() or path.suffix.lower() not in {".png", ".jpg", ".jpeg"}:
                continue
            candidates.append({"path": path, "role": role})

        if not candidates:
            return None, []

        # Strict labeling for image uploads: thumb, body1, body2.
        recent = list(reversed(candidates))

        def pick_role(role: str) -> Optional[Path]:
            for item in recent:
                if item["role"] == role:
                    return item["path"]
            return None

        feature_src = pick_role("thumb")
        detail1_src = pick_role("body1")
        detail2_src = pick_role("body2")

        if feature_src is None and detail1_src is None and detail2_src is None:
            return None, []

        feature_name: Optional[str] = None
        detail_names: list[str] = []

        if feature_src is not None:
            ext = feature_src.suffix.lower()
            target = images_dir / f"feature{ext}"
            shutil.copy2(feature_src, target)
            feature_name = target.name

        if detail1_src is not None:
            ext = detail1_src.suffix.lower()
            target = images_dir / f"detail-1{ext}"
            shutil.copy2(detail1_src, target)
            detail_names.append(target.name)

        if detail2_src is not None:
            ext = detail2_src.suffix.lower()
            target = images_dir / f"detail-2{ext}"
            shutil.copy2(detail2_src, target)
            detail_names.append(target.name)

        return feature_name, detail_names

    def _validate_uploaded_image_roles(self, user_images: list[Dict[str, str]]) -> tuple[bool, list[str]]:
        if not user_images:
            return True, []

        roles = [str(item.get("role", "")).strip().lower() for item in user_images]
        unlabeled = [str(item.get("filename", "")) for item in user_images if not str(item.get("role", "")).strip()]
        missing = [r for r in ["thumb", "body1", "body2"] if r not in roles]

        issues: list[str] = []
        if unlabeled:
            issues.append("Unlabeled images: " + ", ".join(unlabeled))
        if missing:
            issues.append("Missing required labels: " + ", ".join(missing))
        return len(issues) == 0, issues

    def _send_image_naming_help_email(self, draft_id: str, issues: list[str]) -> None:
        draft = self.state["drafts"][draft_id]
        subject = f"[FM-LEAD-DRAFT:{draft_id}] Image naming needed - {draft['title']}"
        body = (
            "I found image attachments, but filenames must include exact role hints for automatic placement.\n\n"
            "Required filename hints:\n"
            "- thumb (featured image)\n"
            "- body1 (after intro)\n"
            "- body2 (before conclusion)\n\n"
            "Examples:\n"
            "- thumb_site.jpg\n"
            "- body1_ops-floor.png\n"
            "- body2_dashboard.jpeg\n\n"
            "Detected issues:\n"
            f"- {'\n- '.join(issues)}\n\n"
            "Please reply again with corrected filenames."
        )
        self._send_email(subject=subject, body=self._with_image_reminder(body))

    def _is_approved(self, cleaned_feedback: str) -> bool:
        for line in cleaned_feedback.splitlines():
            if line.strip().lower() == "approved":
                return True
        return False

    def create_new_draft(self, title: str, brief: str) -> str:
        draft_id = uuid.uuid4().hex[:10]
        slug = self._slugify(title)
        markdown = self._ai_write(self._build_draft_prompt(title=title, brief=brief))
        self.state["drafts"][draft_id] = {
            "id": draft_id,
            "title": title,
            "slug": slug,
            "brief": brief,
            "version": 1,
            "status": "awaiting_feedback",
            "markdown": markdown,
            "created_at": dt.datetime.utcnow().isoformat() + "Z",
        }
        self._save_state()
        attachments = self._create_draft_email_images(
            draft_id=draft_id,
            title=title,
            brief=brief,
            markdown=markdown,
        )
        subject = f"[FM-LEAD-DRAFT:{draft_id}] Focus Mindset Draft v1 - {title}"
        body = (
            "Please review this draft.\n"
            "Reply with corrections, or reply exactly: approved\n"
            "You can attach PDF, DOCX, XLSX files and include URLs as sources.\n\n"
            f"Draft ID: {draft_id}\n"
            f"Title: {title}\n\n"
            f"{markdown}"
        )
        self._send_email(subject=subject, body=self._with_image_reminder(body), attachments=attachments)
        return draft_id

    def _read_unseen_feedback(self) -> None:
        with imaplib.IMAP4_SSL(self.config.imap_host, self.config.imap_port) as mailbox:
            mailbox.login(self.config.imap_username, self.config.imap_password)
            mailbox.select("INBOX")

            queries = [
                f'(FROM "{self.config.owner_email}" SUBJECT "FM-LEAD-DRAFT")',
                f'(FROM "{self.config.owner_email}" SUBJECT "{TOPIC_SUBJECT_PREFIX}")',
            ]
            all_ids = []
            for criteria in queries:
                status, data = mailbox.search(None, criteria)
                if status != "OK" or not data or not data[0]:
                    continue
                all_ids.extend(data[0].split())

            if not all_ids:
                return

            processed = set(self.state.get("processed_message_ids", []))
            # Deduplicate while preserving order.
            msg_ids = list(dict.fromkeys(all_ids))
            # Process newest first to handle the latest editor reply.
            for msg_id in reversed(msg_ids):
                status, fetched = mailbox.fetch(msg_id, "(RFC822)")
                if status != "OK" or not fetched or fetched[0] is None:
                    continue
                raw = fetched[0][1]
                msg = email.message_from_bytes(raw)
                message_id = (msg.get("Message-ID", "") or str(msg_id, errors="ignore")).strip()
                if message_id in processed:
                    continue

                # Ignore agent-authored emails to avoid self-reply loops.
                if str(msg.get("X-FMLead-Agent", "")).strip() == "1":
                    processed.add(message_id)
                    self.state["processed_message_ids"] = list(processed)
                    self._save_state()
                    continue

                subject = msg.get("Subject", "")
                lower_subject = subject.strip().lower()
                is_reply = lower_subject.startswith("re:") or bool(msg.get("In-Reply-To"))
                if not is_reply:
                    continue

                match = DRAFT_TOKEN_RE.search(subject)
                handled = False
                if match:
                    draft_id = match.group(1).lower()
                    if draft_id in self.state["drafts"]:
                        feedback, source_context, user_images = self._extract_feedback_sources(msg=msg, draft_id=draft_id)
                        self._handle_feedback(
                            draft_id=draft_id,
                            feedback=feedback,
                            source_context=source_context,
                            user_images=user_images,
                        )
                        handled = True
                else:
                    topic_title = self._extract_title_from_topic_subject(subject)
                    feedback = self._extract_text_body(msg).strip()
                    topic = self._find_topic_by_title(topic_title)
                    if topic is None:
                        topic = self._find_topic_by_feedback_text(feedback)
                    if topic is not None:
                        self._handle_topic_feedback(topic=topic, feedback=feedback)
                        handled = True

                if not handled:
                    continue

                processed.add(message_id)
                self.state["processed_message_ids"] = list(processed)
                self._save_state()

    def _normalize_feedback(self, feedback: str) -> str:
        lines = []
        for line in feedback.splitlines():
            if line.strip().startswith(">"):
                continue
            if line.strip().lower().startswith("on ") and "wrote:" in line.lower():
                break
            lines.append(line)
        return "\n".join(lines).strip()

    def _handle_feedback(
        self,
        draft_id: str,
        feedback: str,
        source_context: str = "",
        user_images: Optional[list[Dict[str, str]]] = None,
    ) -> None:
        draft = self.state["drafts"][draft_id]
        if str(draft.get("status", "")).strip().lower() == "published":
            return

        cleaned_feedback = self._normalize_feedback(feedback)
        if not cleaned_feedback:
            return

        parsed_labels = self._extract_image_labels_from_feedback(cleaned_feedback)
        if parsed_labels:
            current_labels = draft.get("image_labels", {})
            if not isinstance(current_labels, dict):
                current_labels = {}
            current_labels.update(parsed_labels)
            draft["image_labels"] = current_labels

        if user_images:
            existing = draft.get("user_image_attachments", [])
            if not isinstance(existing, list):
                existing = []
            existing.extend(user_images)
            # Keep only recent history to avoid unbounded growth.
            draft["user_image_attachments"] = existing[-30:]

        valid_images, issues = self._validate_uploaded_image_roles(user_images or [])

        if self._is_approved(cleaned_feedback):
            if (user_images or []) and not valid_images:
                draft["status"] = "awaiting_feedback"
                self._save_state()
                self._send_image_naming_help_email(draft_id=draft_id, issues=issues)
                return
            draft["status"] = "approved"
            self._save_state()
            self.publish(draft_id)
            return

        if source_context.strip():
            updates = draft.get("source_updates", [])
            updates.append({"at": dt.datetime.utcnow().isoformat() + "Z", "chars": len(source_context)})
            draft["source_updates"] = updates

        revised = self._ai_write(
            self._build_revision_prompt(
                current_markdown=draft["markdown"],
                feedback=cleaned_feedback,
                source_context=source_context,
            )
        )
        draft["markdown"] = revised
        draft["version"] = int(draft.get("version", 1)) + 1
        draft["status"] = "awaiting_feedback"
        self._save_state()

        subject = f"[FM-LEAD-DRAFT:{draft_id}] Focus Mindset Draft v{draft['version']} - {draft['title']}"
        body = (
            "Updated draft based on your feedback.\n"
            "Reply with more corrections, or reply exactly: approved\n"
            "You can attach PDF, DOCX, XLSX files and include URLs as sources.\n\n"
            f"Draft ID: {draft_id}\n"
            f"Version: {draft['version']}\n\n"
            f"{draft['markdown']}"
        )
        attachments = self._create_draft_email_images(
            draft_id=draft_id,
            title=draft["title"],
            brief=draft.get("brief", ""),
            markdown=draft["markdown"],
        )
        self._send_email(subject=subject, body=self._with_image_reminder(body), attachments=attachments)

    def publish(self, draft_id: str) -> None:
        draft = self.state["drafts"][draft_id]
        if draft.get("status") not in {"approved", "awaiting_feedback"}:
            return
        slug = draft["slug"]
        post_dir = self._article_root() / slug
        images_dir = post_dir / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        (images_dir / ".gitkeep").write_text("", encoding="utf-8")
        feature_image_name, detail_image_names = self._select_user_publish_images(draft=draft, images_dir=images_dir)

        image_queries: Dict[str, str] = {}
        if not feature_image_name and not detail_image_names:
            if self.config.image_query_mode == "source":
                image_queries = self._suggest_image_queries(
                    title=draft["title"],
                    brief=draft.get("brief", ""),
                    markdown=draft.get("markdown", ""),
                )

            feature_image_name = self._generate_feature_image(
                title=draft["title"],
                brief=draft.get("brief", ""),
                images_dir=images_dir,
                suggested_query=image_queries.get("feature"),
            )
            detail_image_names = self._generate_inpost_images(
                title=draft["title"],
                brief=draft.get("brief", ""),
                images_dir=images_dir,
                suggested_queries=image_queries,
            )

        now = dt.datetime.now().strftime("%Y-%m-%d")
        tags = [t.strip() for t in self.config.default_tags.split(",") if t.strip()]
        tags_yaml = "\n".join([f"  - \"{t}\"" for t in tags])
        image_labels = draft.get("image_labels", {}) if isinstance(draft.get("image_labels", {}), dict) else {}
        thumb_label = str(image_labels.get("thumb", "")).replace('"', "'").strip()
        feature_line = f"featureimage: \"images/{feature_image_name}\"\n" if feature_image_name else ""
        feature_alt_line = f"featureimage_alt: \"{thumb_label}\"\n" if feature_image_name and thumb_label else ""
        front_matter = (
            "---\n"
            f"title: \"{draft['title']}\"\n"
            f"date: {now}\n"
            "draft: false\n"
            "description: \"Focus Mindset article\"\n"
            f"{feature_line}"
            f"{feature_alt_line}"
            "tags:\n"
            f"{tags_yaml}\n"
            "---\n\n"
        )
        detail_image_markdown = ""
        for idx, name in enumerate(detail_image_names, start=1):
            role = f"body{idx}"
            label = str(image_labels.get(role, "")).replace("\n", " ").strip()
            alt_text = label or f"Supporting image {idx} for {draft['title']}"
            detail_image_markdown += f"![{alt_text}](images/{name})\n"
            if label:
                detail_image_markdown += f"*{label}*\n"
            detail_image_markdown += "\n"
        post_dir.mkdir(parents=True, exist_ok=True)
        (post_dir / "index.md").write_text(front_matter + detail_image_markdown + draft["markdown"].strip() + "\n", encoding="utf-8")

        self._run_cmd(["hugo", "--renderToMemory"], cwd=self.config.hugo_site_path)
        self._run_cmd(["git", "add", str(post_dir.relative_to(self.config.hugo_site_path))], cwd=self.config.hugo_site_path)
        self._run_cmd(["git", "commit", "-m", f"Add Focus Mindset article: {draft['title']}"], cwd=self.config.hugo_site_path)
        self._run_cmd(["git", "push"], cwd=self.config.hugo_site_path)

        draft["status"] = "published"
        draft["published_at"] = dt.datetime.utcnow().isoformat() + "Z"
        self._save_state()
        subject = f"[FM-LEAD-DRAFT:{draft_id}] Pushed to GitHub - {draft['title']}"
        body = (
            "Article approved and committed/pushed to GitHub.\n"
            "Note: this email confirms Git push only (not live site deployment status).\n\n"
            f"Path: {post_dir.as_posix()}/index.md\n"
            f"Images: {images_dir.as_posix()}\n"
            f"Feature image: {feature_image_name or 'not generated'}\n"
            f"In-post images: {', '.join(detail_image_names) if detail_image_names else 'not generated'}\n"
            f"Image query mode: {self.config.image_query_mode}\n"
            f"Queries: {json.dumps(image_queries) if image_queries else 'default topic queries'}"
        )
        self._send_email(subject=subject, body=body)

    def watch(self) -> None:
        if not self._acquire_watch_lock():
            print("watch already running: another watcher instance holds the singleton lock")
            return
        while True:
            try:
                self._maybe_send_daily_topics()
                self._maybe_send_weekly_topic_reminder()
                self._read_unseen_feedback()
            except Exception as exc:
                print(f"watch loop error: {exc}")
            time.sleep(self.config.poll_seconds)

    def _acquire_watch_lock(self) -> bool:
        # Hard-coded singleton guard: only one watch loop can bind this local port.
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            sock.bind(("127.0.0.1", WATCH_SINGLETON_PORT))
            sock.listen(1)
            self._watch_lock_socket = sock
            return True
        except OSError:
            sock.close()
            return False

    @staticmethod
    def _run_cmd(cmd, cwd: Path) -> None:
        subprocess.run(cmd, cwd=str(cwd), check=True)


def load_config(env_path: Path) -> Config:
    load_dotenv(dotenv_path=env_path)

    def require(name: str) -> str:
        value = os.getenv(name, "").strip()
        if not value:
            raise ValueError(f"Missing required env var: {name}")
        return value

    def parse_bool(name: str, default: bool) -> bool:
        raw = os.getenv(name)
        if raw is None:
            return default
        return raw.strip().lower() in {"1", "true", "yes", "on"}

    llm_backend = os.getenv("LLM_BACKEND", "openai").strip().lower()
    openai_key = os.getenv("OPENAI_API_KEY", "").strip()
    if llm_backend == "openai" and not openai_key:
        raise ValueError("Missing required env var: OPENAI_API_KEY")

    return Config(
        openai_api_key=openai_key,
        openai_model=os.getenv("OPENAI_MODEL", "gpt-5.3-codex"),
        smtp_host=require("SMTP_HOST"),
        smtp_port=int(os.getenv("SMTP_PORT", "465")),
        smtp_username=require("SMTP_USERNAME"),
        smtp_password=require("SMTP_PASSWORD"),
        imap_host=require("IMAP_HOST"),
        imap_port=int(os.getenv("IMAP_PORT", "993")),
        imap_username=require("IMAP_USERNAME"),
        imap_password=require("IMAP_PASSWORD"),
        owner_email=require("OWNER_EMAIL"),
        hugo_site_path=Path(require("HUGO_SITE_PATH")),
        article_root_relative=os.getenv("ARTICLE_ROOT_RELATIVE", "content/focus/articles"),
        default_tags=os.getenv("DEFAULT_TAGS", "focus mindset,leadership"),
        poll_seconds=int(os.getenv("POLL_SECONDS", "30")),
        image_model=os.getenv("IMAGE_MODEL", "gpt-image-1"),
        image_size=os.getenv("IMAGE_SIZE", "1536x1024"),
        image_quality=os.getenv("IMAGE_QUALITY", "high"),
        image_style=os.getenv("IMAGE_STYLE", "natural"),
        image_prompt_suffix=os.getenv("IMAGE_PROMPT_SUFFIX", ""),
        image_provider=os.getenv("IMAGE_PROVIDER", "unsplash").strip().lower(),
        image_query_mode=os.getenv("IMAGE_QUERY_MODE", "source").strip().lower(),
        generate_detail_image=parse_bool("GENERATE_DETAIL_IMAGE", True),
        in_post_image_count=int(os.getenv("IN_POST_IMAGE_COUNT", "2")),
        llm_backend=llm_backend,
        ollama_url=os.getenv("OLLAMA_URL", "http://127.0.0.1:11434"),
        ollama_model=os.getenv("OLLAMA_MODEL", "llama3.1:8b"),
        ollama_timeout_seconds=int(os.getenv("OLLAMA_TIMEOUT_SECONDS", "600")),
        max_source_chars=int(os.getenv("MAX_SOURCE_CHARS", "12000")),
        enable_url_fetch=parse_bool("ENABLE_URL_FETCH", True),
        generate_images=parse_bool("GENERATE_IMAGES", True),
        tracker_filename=os.getenv("TRACKER_FILENAME", "fmlead_post_tracker.xlsx"),
        daily_topics_enabled=parse_bool("DAILY_TOPICS_ENABLED", True),
        daily_topics_count=int(os.getenv("DAILY_TOPICS_COUNT", "3")),
        daily_topics_hour_24=int(os.getenv("DAILY_TOPICS_HOUR_24", "9")),
        idea_pipeline_max=max(1, int(os.getenv("IDEA_PIPELINE_MAX", "10"))),
        weekly_idea_reminder_enabled=parse_bool("WEEKLY_IDEA_REMINDER_ENABLED", True),
        weekly_reminder_day_0_mon=max(0, min(6, int(os.getenv("WEEKLY_REMINDER_DAY_0_MON", "0")))),
        weekly_reminder_hour_24=max(0, min(23, int(os.getenv("WEEKLY_REMINDER_HOUR_24", "9")))),
        attach_images_in_draft_email=parse_bool("ATTACH_IMAGES_IN_DRAFT_EMAIL", True),
        image_thumb_filename=os.getenv("IMAGE_THUMB_FILENAME", "THUMB_test.jpg").strip() or "THUMB_test.jpg",
        image_body1_filename=os.getenv("IMAGE_BODY1_FILENAME", "BODY1_test.jpg").strip() or "BODY1_test.jpg",
        image_body2_filename=os.getenv("IMAGE_BODY2_FILENAME", "BODY2_test.jpg").strip() or "BODY2_test.jpg",
    )


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description=f"{AGENT_NAME} email content agent")
    parser.add_argument("command", choices=["new", "watch", "publish", "daily-topics"], help="Run one command")
    parser.add_argument("--title", help="Draft title")
    parser.add_argument("--brief", help="Draft brief/instructions")
    parser.add_argument("--draft-id", help="Draft id for publish")
    parser.add_argument("--env-file", default=str(script_dir / ".env"), help="Path to env file")

    args = parser.parse_args()
    config = load_config(Path(args.env_file))
    state_path = script_dir / "state.json"
    agent = FMLeadComFocusAWriter(config=config, state_path=state_path)

    if args.command == "new":
        if not args.title or not args.brief:
            raise ValueError("new command requires --title and --brief")
        draft_id = agent.create_new_draft(title=args.title, brief=args.brief)
        print(f"Created and emailed draft: {draft_id}")
        return
    if args.command == "watch":
        agent.watch()
        return
    if args.command == "publish":
        if not args.draft_id:
            raise ValueError("publish command requires --draft-id")
        agent.publish(args.draft_id)
        print(f"Published draft: {args.draft_id}")
        return

    if args.command == "daily-topics":
        sent = agent.send_daily_topic_approval_requests()
        print(f"Sent topic approval emails: {sent}")
        return


if __name__ == "__main__":
    main()


# Backward-compatible alias for older scripts importing the previous class name.
ContentEmailAgent = FMLeadComFocusAWriter
